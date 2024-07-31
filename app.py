from flask import Flask, render_template, request, redirect, url_for, session
from pymongo import MongoClient
from werkzeug.security import generate_password_hash, check_password_hash
import configparser
import openpyxl
import json
from config.config_reader import load_config
from auth import auth_bp  # Import the auth Blueprint

config = load_config()
mongo_uri = config['MONGO_URI']
database_name = config['DATABASE_NAME']
collection_name = config['COLLECTION_NAME']
secret_key = config['SECRET_KEY']

app = Flask(__name__)
app.secret_key = secret_key

client = MongoClient(mongo_uri)
db = client[database_name]
users_collection = db[collection_name]

app.register_blueprint(auth_bp)  # Register the auth Blueprint

# Load the sheet names from the Excel file
def get_test_names(filename):
    workbook = openpyxl.load_workbook(filename, data_only=True)
    return [sheet for sheet in workbook.sheetnames if sheet.startswith("TEST")]

# Load the questions from the Excel file
def load_questions(filename, sheet_name):
    workbook = openpyxl.load_workbook(filename, data_only=True)
    sheet = workbook[sheet_name]
    questions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        question_text = row[0]
        answer_text = row[1]
        if question_text and answer_text:  # Check if both question and answer are present
            questions.append({'question': question_text, 'answer': answer_text})
    return questions

@app.route('/')
def select_test():
    if 'student_id' not in session:
        return redirect(url_for('auth.login'))
    test_names = get_test_names('questions.xlsx')
    return render_template('select_test.html', test_names=test_names)

@app.route('/test/<test_name>', methods=['GET', 'POST'])
def test(test_name):
    if 'student_id' not in session:
        return redirect(url_for('auth.login'))
    
    questions = load_questions('questions.xlsx', test_name)
    results = {}
    retry_mode = False  # Default to False

    if request.method == 'POST':
        user_answers = request.form.to_dict()
        incorrect_count = 0

        for question in questions:
            question_text = question['question']
            user_answer = user_answers.get(question_text, '').strip()
            correct_answer = str(question['answer']).strip()

            is_correct = user_answer.lower() == correct_answer.lower()
            if not is_correct:
                incorrect_count += 1

            results[question_text] = {
                'user_answer': user_answer,
                'is_correct': is_correct,
                'correct_answer': correct_answer
            }

        if incorrect_count == 0:
            return redirect(url_for('congratulations'))

        # Redirect with JSON encoded user answers
        return redirect(url_for('error_page', incorrect_count=incorrect_count, test_name=test_name, user_answers=json.dumps(user_answers)))

    # For GET request or retry scenario
    user_answers = request.args.get('user_answers', '{}')
    retry_mode = 'retry' in request.args
    results = {}

    user_answers_dict = json.loads(user_answers) if user_answers else {}

    for question in questions:
        question_text = question['question']
        user_answer = user_answers_dict.get(question_text, '').strip()
        correct_answer = str(question['answer']).strip()

        is_correct = user_answer.lower() == correct_answer.lower()
        results[question_text] = {
            'user_answer': user_answer,
            'is_correct': is_correct,
            'correct_answer': correct_answer
        }

    return render_template('index.html', questions=questions, test_name=test_name, results=results, retry_mode=retry_mode, user_answers=user_answers_dict)

@app.route('/error_page')
def error_page():
    if 'student_id' not in session:
        return redirect(url_for('auth.login'))

    incorrect_count = request.args.get('incorrect_count', 0, type=int)
    test_name = request.args.get('test_name', '')
    user_answers = request.args.get('user_answers', '{}')

    # Pass JSON encoded user_answers
    return render_template('error_page.html', incorrect_count=incorrect_count, test_name=test_name, user_answers=user_answers)

@app.route('/congratulations')
def congratulations():
    if 'student_id' not in session:
        return redirect(url_for('auth.login'))
    return render_template('congratulations.html')

if __name__ == '__main__':
    app.run(debug=True)
